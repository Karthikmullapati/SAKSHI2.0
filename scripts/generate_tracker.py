import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

# Load the master template workbook
src_file = 'Sakshi_AI_Gen3_Project_Plan.xlsx'
dst_file = 'Sakshi_AI_Gen3_Team_Contribution_Tracker.xlsx'

wb = openpyxl.load_workbook(src_file)

# Common Styling Definitions
font_title = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
fill_title = PatternFill(start_color='0B2545', end_color='0B2545', fill_type='solid')

font_section = Font(name='Calibri', size=12, bold=True, color='0B2545')

font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

font_data = Font(name='Calibri', size=10, bold=False, color='000000')
font_data_bold = Font(name='Calibri', size=10, bold=True, color='000000')

# Status Fills & Fonts
fill_done = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
font_done = Font(name='Calibri', size=10, color='006100', bold=True)

fill_inprogress = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
font_inprogress = Font(name='Calibri', size=10, color='9C6500', bold=True)

fill_partial = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
font_partial = Font(name='Calibri', size=10, color='B25900', bold=True)

fill_pending = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
font_pending = Font(name='Calibri', size=10, color='595959', bold=False)

fill_verification = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
font_verification = Font(name='Calibri', size=10, color='203764', bold=True)

fill_blocked = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
font_blocked = Font(name='Calibri', size=10, color='9C0006', bold=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')


def clear_sheet_contents(ws):
    # Unmerge all merged ranges first to avoid MergedCell read-only errors
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))
    
    for r in range(1, ws.max_row + 50):
        for c in range(1, ws.max_column + 15):
            cell = ws.cell(r, c)
            try:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
            except Exception:
                pass


def apply_status_style(cell, status_val):
    val = str(status_val).strip()
    if val in ('Done', 'Completed', 'Passed'):
        cell.fill = fill_done
        cell.font = font_done
    elif val in ('In Progress', 'Active'):
        cell.fill = fill_inprogress
        cell.font = font_inprogress
    elif val in ('Partially Completed', 'Partial'):
        cell.fill = fill_partial
        cell.font = font_partial
    elif val in ('Needs Verification', 'Needs Manual Verification', 'Not Yet Automated'):
        cell.fill = fill_verification
        cell.font = font_verification
    elif val in ('Blocked', 'Failed'):
        cell.fill = fill_blocked
        cell.font = font_blocked
    elif val in ('Pending', 'Not Started', 'Not Implemented'):
        cell.fill = fill_pending
        cell.font = font_pending
    else:
        cell.font = font_data


# ==========================================
# 1. SHEET: Executive Summary
# ==========================================
ws_exec = wb['Executive Summary']
ws_exec.views.sheetView[0].showGridLines = True
clear_sheet_contents(ws_exec)

# Title
ws_exec.cell(1, 1, 'Sakshi AI Gen 3 — Team Contribution Tracker & Master Project Status')
ws_exec.merge_cells('A1:E1')
ws_exec.cell(1, 1).font = font_title
ws_exec.cell(1, 1).fill = fill_title
ws_exec.cell(1, 1).alignment = align_left
ws_exec.row_dimensions[1].height = 35

# Project Overview Section
ws_exec.cell(3, 1, 'PROJECT OVERVIEW')
ws_exec.cell(3, 1).font = font_section

overview_headers = ['Field', 'Details']
for c, h in enumerate(overview_headers, 1):
    cell = ws_exec.cell(4, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_left
    cell.border = thin_border
ws_exec.row_dimensions[4].height = 24

overview_data = [
    ('Project Name', 'Sakshi AI Gen 3 — Enterprise Multi-Tenant Invoice, Tax & GL Automation Engine'),
    ('Architecture & Core Engineering', 'Rahul (Primary Lead) & Abhishek (Core Engineering Co-Lead)'),
    ('Compliance & ERP Integration Lead', 'Raju (TDS Withholding Brain, GST Rules & Zoho Books OAuth2/Export)'),
    ('Channel Ingestion & Staging Lead', 'Neeraj (IMAP Multi-Server Email Ingestion, PDF Renderer & Inbox Pipeline)'),
    ('Current Project Health', '81.8% Core Functional Readiness | 111 Passed Automated Tests | Live Staging Active'),
    ('Primary Business Deliverables', 'AI Extraction, 7-Pillar HITL Review, Deterministic GST/ITC/TDS, Balanced GL & Direct Zoho Sync'),
    ('Current Active Focus', 'Resolution of Unbalanced Extraction Edge Cases, Full Multi-State E2E UAT, Production Deployment')
]

for r_idx, (k, v) in enumerate(overview_data, 5):
    c1 = ws_exec.cell(r_idx, 1, k)
    c2 = ws_exec.cell(r_idx, 2, v)
    c1.font = font_data_bold
    c1.border = thin_border
    c1.alignment = align_left
    c2.font = font_data
    c2.border = thin_border
    c2.alignment = align_left
    ws_exec.row_dimensions[r_idx].height = 20

# Team Core Ownership Summary Section
r_start_team = 13
ws_exec.cell(r_start_team, 1, 'TEAM FUNCTIONAL OWNERSHIP & WORK ALLOCATION')
ws_exec.cell(r_start_team, 1).font = font_section

team_headers = ['Team Member', 'Primary Functional Ownership', 'Secondary / Support Areas', 'Status']
for c, h in enumerate(team_headers, 1):
    cell = ws_exec.cell(r_start_team + 1, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_left
    cell.border = thin_border
ws_exec.row_dimensions[r_start_team + 1].height = 24

team_ownership_rows = [
    ('Rahul', 'Backend Core Architecture, DB Models, Duplicate Detection, Financial Validation Engine, Security/Auth RBAC', 'Frontend Invoices Hub, HITL Review UI, Automated Test Suite', 'Active'),
    ('Abhishek', 'Frontend UI Architecture, 7-Pillar Review Workspace, GL Journal Inspector, Accounting Period & Multi-Tenant Controls', 'Backend API Routers, Invoice Pipeline, Storage Service, E2E Testing', 'Active'),
    ('Raju', 'TDS Withholding Engine (Sec 194C/J/I/Q), Zoho Books OAuth2 Client, Master Data Sync & GL Bill Export Service', 'GST & ITC Compliance Rules, Accounting Classification Validation', 'Active'),
    ('Neeraj', 'IMAP Multi-Server Email Ingestion, Background Polling Service, PDF Multi-Page Renderer, Staged Inbox Workspace', 'Settings Configuration UI, Duplicate Attachment Filtering', 'Active')
]

for r_idx, row in enumerate(team_ownership_rows, r_start_team + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_exec.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx == 1 else font_data
        cell.border = thin_border
        cell.alignment = align_center if c_idx == 4 else align_left
        if c_idx == 4:
            apply_status_style(cell, val)
    ws_exec.row_dimensions[r_idx].height = 20

# Key Milestones Schedule
r_start_ms = 19
ws_exec.cell(r_start_ms, 1, 'KEY MILESTONES SCHEDULE & ACTUAL STATUS')
ws_exec.cell(r_start_ms, 1).font = font_section

ms_headers = ['Milestone ID', 'Target Schedule', 'Deliverable Description', 'Primary Owner', 'Actual Status']
for c, h in enumerate(ms_headers, 1):
    cell = ws_exec.cell(r_start_ms + 1, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_left
    cell.border = thin_border
ws_exec.row_dimensions[r_start_ms + 1].height = 24

milestones_data = [
    ('MS-01', 'Phase 1', 'Storage, Multi-Tenant Database, Schema Validation & Qwen-VL Integration', 'Rahul (Primary) / Abhishek (Support)', 'Completed'),
    ('MS-02', 'Phase 2', 'Deterministic GST, ITC, TDS & Stage 5 Financial Validation Engines', 'Raju (TDS) / Rahul (Validation) / Abhishek (GST/ITC)', 'Completed'),
    ('MS-03', 'Phase 3', 'Double-Entry Journal Generator, Relational GL Ledger & HITL Approval Lock', 'Abhishek (Primary) / Rahul (Support)', 'Completed'),
    ('MS-04', 'Phase 3', 'Resilient Zoho Books OAuth2 Middleware, Master Data Sync & Bill Export', 'Raju', 'Completed'),
    ('MS-05', 'Phase 3', 'IMAP Email Auto-Ingestion Engine, PDF Page Renderer & Staged Inbox UI', 'Neeraj', 'Completed'),
    ('MS-06', 'Phase 4', 'Frontend 7-Pillar Review Hub, GL Inspector & Integrations Management', 'Abhishek (Primary) / Rahul (Support)', 'Completed'),
    ('MS-07', 'Phase 4', 'Comprehensive Automated Test Suite (111 Backend Unit & Integration Tests)', 'Rahul / Abhishek / Raju / Neeraj', 'Completed'),
    ('MS-08', 'Phase 5', 'End-to-End Enterprise UAT Approval, Production Hardening & Cloud Rollout', 'Rahul / Abhishek / Raju / Neeraj', 'In Progress')
]

for r_idx, row in enumerate(milestones_data, r_start_ms + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_exec.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx in (1, 4) else font_data
        cell.border = thin_border
        cell.alignment = align_center if c_idx in (1, 2, 5) else align_left
        if c_idx == 5:
            apply_status_style(cell, val)
    ws_exec.row_dimensions[r_idx].height = 20

# High-Level Metrics Summary Box
r_start_metrics = 29
ws_exec.cell(r_start_metrics, 1, 'CURRENT PROJECT EXECUTION METRICS')
ws_exec.cell(r_start_metrics, 1).font = font_section

metric_headers = ['Metric Category', 'Total Inventory', 'Completed / Verified', 'In Progress / Partial', 'Pending / Next Up']
for c, h in enumerate(metric_headers, 1):
    cell = ws_exec.cell(r_start_metrics + 1, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_left
    cell.border = thin_border
ws_exec.row_dimensions[r_start_metrics + 1].height = 24

metric_rows = [
    ('Master WBS Tasks (Total Inventory: 44 Tasks)', '44 Tasks', '36 Done (81.8%)', '5 In Progress / Partial (11.4%)', '3 Pending (6.8%)'),
    ('QA Automated & Integration Tests', '113 Test Cases', '111 Passed (98.2%)', '2 Needs PyMuPDF / Async Fix', '0 Failed'),
    ('Enterprise Functional Use Cases', '43 Scenarios', '34 Verified / Active', '5 In Progress / Verification', '4 Future / Edge Scenarios')
]

for r_idx, row in enumerate(metric_rows, r_start_metrics + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_exec.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx == 1 else font_data
        cell.border = thin_border
        cell.alignment = align_left if c_idx == 1 else align_center
    ws_exec.row_dimensions[r_idx].height = 20


# ==========================================
# 2. SHEET: Master Task List (WBS)
# ==========================================
ws_wbs = wb['Master Task List (WBS)']
ws_wbs.views.sheetView[0].showGridLines = True
clear_sheet_contents(ws_wbs)

# Title
ws_wbs.cell(1, 1, 'Master Task List & Work Breakdown Structure (WBS) — Actual Repository Implementation')
ws_wbs.merge_cells('A1:H1')
ws_wbs.cell(1, 1).font = font_title
ws_wbs.cell(1, 1).fill = fill_title
ws_wbs.cell(1, 1).alignment = align_left
ws_wbs.row_dimensions[1].height = 35

# Table Headers
wbs_headers = ['Task ID', 'Module / Component', 'Task Description', 'Primary Owner', 'Est. Hours', 'Start Date', 'End Date', 'Status']
for c, h in enumerate(wbs_headers, 1):
    cell = ws_wbs.cell(3, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center if c in (1, 5, 6, 7, 8) else align_left
    cell.border = thin_border
ws_wbs.row_dimensions[3].height = 26

# Comprehensive Task Inventory from Repository
# Rahul / Abhishek balanced roughly equal (14 Primary each)
# Raju owns TDS & Zoho (9 Tasks)
# Neeraj owns Email / IMAP & Staging (6 Tasks)
# All Team Members for UAT (1 Task)
wbs_data = [
    # 1. Ingestion & Storage
    ('WBS 1.1', 'Invoice Ingestion & Storage', 'Multi-tenant file upload API with SHA-256 deduplication and local disk storage backend', 'Rahul (Primary) / Abhishek (Support)', 20.0, '2026-08-01', '2026-08-03', 'Done'),
    ('WBS 1.2', 'Duplicate Detection', 'Hash-based and fuzzy vendor/invoice_no/date duplicate prevention and staging promotion', 'Abhishek (Primary) / Rahul (Support)', 16.0, '2026-08-03', '2026-08-05', 'Done'),
    ('WBS 1.3', 'IMAP Email Ingestion Service', 'Background email polling engine with SSL/TLS, multi-server support, and attachment extraction', 'Neeraj', 24.0, '2026-08-10', '2026-08-14', 'Done'),
    ('WBS 1.4', 'Inbox Staging API', 'Staged document lifecycle management, duplicate filtering, and batch manual promotion to processing', 'Neeraj', 18.0, '2026-08-14', '2026-08-18', 'Done'),
    ('WBS 1.5', 'PDF Multi-Page Renderer', 'PyMuPDF-based multi-page PDF document inspection and base64 thumbnail rendering service', 'Neeraj', 16.0, '2026-08-18', '2026-08-21', 'Done'),
    ('WBS 1.6', 'IMAP Configuration & Settings API', 'Secure storage of user IMAP credentials with encryption, connection test endpoint, and disconnect handler', 'Neeraj', 14.0, '2026-08-21', '2026-08-24', 'Done'),

    # 2. AI Extraction & Normalization
    ('WBS 2.1', 'Qwen-VL Vision Client', 'FastAPI integration with Qwen3-VL Colab GPU server for Indian B2B invoice OCR & key-value parsing', 'Rahul (Primary) / Abhishek (Support)', 24.0, '2026-08-02', '2026-08-06', 'Done'),
    ('WBS 2.2', 'Pydantic Extraction Schemas', 'Comprehensive validation schemas for invoice headers, line items, bank details, and PAN/GSTIN slicers', 'Rahul (Primary) / Abhishek (Support)', 18.0, '2026-08-05', '2026-08-08', 'Done'),
    ('WBS 2.3', 'Dual-State JSONB Architecture', 'Raw vs Current VLM and Accounting outputs preservation for full HITL audit trail and rollback', 'Abhishek (Primary) / Rahul (Support)', 16.0, '2026-08-08', '2026-08-11', 'Done'),
    ('WBS 2.4', 'Invoice Data Normalization', 'Dynamic numeric parser, Indian date normalizer, and multi-currency symbol cleaner', 'Rahul (Primary) / Abhishek (Support)', 14.0, '2026-08-10', '2026-08-12', 'Done'),
    ('WBS 2.5', 'Non-Invoice Document Detection', 'AI rejection rules for non-invoices (statements, receipts, IDs, delivery challans)', 'Abhishek (Primary) / Rahul (Support)', 12.0, '2026-08-12', '2026-08-14', 'Done'),

    # 3. Chart of Accounts & TDS Brain
    ('WBS 3.1', 'COA Classification Engine', 'Qwen3-4B semantic keyword & rule-based Chart of Accounts categorization for line items', 'Abhishek (Primary) / Rahul (Support)', 20.0, '2026-08-12', '2026-08-16', 'Done'),
    ('WBS 3.2', 'TDS Withholding Engine', 'Deterministic Section 194C, 194J (2%/10%), 194I (Rent), 194Q & 206AB threshold calculation', 'Raju', 24.0, '2026-08-14', '2026-08-18', 'Done'),
    ('WBS 3.3', 'Vendor PAN & TDS Rules Slicer', 'PAN 4th character entity detector (Company vs Individual) and PAN non-furnishing 20% penalty rules', 'Raju', 16.0, '2026-08-18', '2026-08-20', 'Done'),
    ('WBS 3.4', 'Cost Centre & Dimensions Engine', 'Tracking cost centers, project codes, and departments across line items', 'Abhishek (Primary) / Rahul (Support)', 14.0, '2026-08-20', '2026-08-22', 'Partially Completed'),

    # 4. GST & ITC Compliance
    ('WBS 4.1', 'Deterministic GST Engine', 'Intra-State vs Inter-State supply type determination, state code resolver, and tax rate matcher', 'Abhishek (Primary) / Rahul (Support)', 20.0, '2026-08-15', '2026-08-19', 'Done'),
    ('WBS 4.2', 'SEZ & RCM Tax Logic', 'Special Economic Zone 0% IGST overrides and Reverse Charge Mechanism tax liability handler', 'Rahul (Primary) / Abhishek (Support)', 16.0, '2026-08-19', '2026-08-21', 'Done'),
    ('WBS 4.3', 'ITC Eligibility Engine', 'Section 16 eligibility criteria evaluation and Section 17(5) blocked credit classifier', 'Abhishek (Primary) / Rahul (Support)', 18.0, '2026-08-20', '2026-08-23', 'Done'),

    # 5. Financial Validation & Journal Generation
    ('WBS 5.1', 'Financial Validation Engine', 'Independent mathematical cross-reconciliation (lines, subtotal, tax breakdown, grand total)', 'Rahul (Primary) / Abhishek (Support)', 22.0, '2026-08-21', '2026-08-24', 'Done'),
    ('WBS 5.2', 'Reconciliation Anomaly Detector', 'Line-item math mismatch detection, rounding difference handler, and discrepancy warnings', 'Rahul (Primary) / Abhishek (Support)', 16.0, '2026-08-23', '2026-08-25', 'Done'),
    ('WBS 5.3', 'Double-Entry Journal Generator', 'Deterministic balanced debits (expenses/assets/input tax) vs credits (AP/TDS) preview generator', 'Abhishek (Primary) / Rahul (Support)', 24.0, '2026-08-22', '2026-08-25', 'Done'),
    ('WBS 5.4', 'Relational GL Database Sync', 'Idempotent persistence to relational journal_entries and journal_lines tables', 'Abhishek (Primary) / Rahul (Support)', 18.0, '2026-08-24', '2026-08-26', 'Done'),
    ('WBS 5.5', 'Authoritative Accounting Validation', 'Mandatory Finance approval verification on COA lines and zero-fallback strict journal lock', 'Rahul (Primary) / Abhishek (Support)', 16.0, '2026-08-25', '2026-08-27', 'Done'),

    # 6. HITL Review, Approval & Audit
    ('WBS 6.1', 'Invoice Review API', 'CRUD operations for draft edits, line-item account overrides, and draft journal preview', 'Abhishek (Primary) / Rahul (Support)', 18.0, '2026-08-23', '2026-08-26', 'Done'),
    ('WBS 6.2', 'Strict Approval & Lock Workflow', 'Atomic invoice approval state machine, balanced journal gatekeeper, and tamper-proof lock', 'Rahul (Primary) / Abhishek (Support)', 18.0, '2026-08-26', '2026-08-28', 'Done'),
    ('WBS 6.3', 'Rejection & Reason Logging', 'Invoice rejection workflow with mandatory audit trail reason logging', 'Abhishek (Primary) / Rahul (Support)', 12.0, '2026-08-26', '2026-08-28', 'Done'),
    ('WBS 6.4', 'Audit Trail Service', 'Detailed event logging for all user interactions, state mutations, and export attempts', 'Rahul (Primary) / Abhishek (Support)', 14.0, '2026-08-26', '2026-08-28', 'Done'),
    ('WBS 6.5', 'Accounting Period Controls', 'Fiscal calendar validation, closed period locking, and back-dated posting controls', 'Rahul (Primary) / Abhishek (Support)', 16.0, '2026-08-28', '2026-08-30', 'Partially Completed'),

    # 7. Zoho Books ERP Integration
    ('WBS 7.1', 'Zoho OAuth2 Client', 'Resilient OAuth2 token lifecycle management with automated refresh and secure encryption', 'Raju', 24.0, '2026-08-16', '2026-08-20', 'Done'),
    ('WBS 7.2', 'Zoho Master Data Sync', 'Synchronizing Chart of Accounts, Tax Rates, Vendors, Payment Terms, and Cost Centers', 'Raju', 20.0, '2026-08-20', '2026-08-24', 'Done'),
    ('WBS 7.3', 'Zoho Vendor Auto-Upsert', 'Auto-creation and update of vendor contacts in Zoho with GSTIN, PAN, and bank details', 'Raju', 18.0, '2026-08-22', '2026-08-25', 'Done'),
    ('WBS 7.4', 'Zoho Bill Export Service', 'Direct export of approved invoices as Zoho Bills with line accounts, tax mappings, and TDS credits', 'Raju', 22.0, '2026-08-24', '2026-08-28', 'Done'),
    ('WBS 7.5', 'Zoho Error Handling & Auto-Healer', 'Resilient error recovery for rate limits, effective date lock retries, and duplicate bill numbers', 'Raju', 16.0, '2026-08-26', '2026-08-29', 'Done'),
    ('WBS 7.6', 'Credit / Debit Notes Support', 'Handling vendor credit notes and debit adjustments in Zoho sync workflow', 'Raju', 16.0, '2026-08-28', '2026-08-31', 'In Progress'),

    # 8. Frontend Review & Operations Hub
    ('WBS 8.1', 'Next.js App Architecture', 'AppShell navigation, responsive layout, sidebar routes, and authenticated state handling', 'Abhishek (Primary) / Rahul (Support)', 20.0, '2026-08-18', '2026-08-22', 'Done'),
    ('WBS 8.2', '7-Pillar Review Workspace', 'Interactive split-view review page (PDF/Image viewer + 7 validation and accounting pillars)', 'Abhishek (Primary) / Rahul (Support)', 28.0, '2026-08-22', '2026-08-26', 'Done'),
    ('WBS 8.3', 'Journal & GL Inspector UI', 'Live double-entry preview card showing Debits vs Credits, balance status badge, and diffs', 'Abhishek (Primary) / Rahul (Support)', 16.0, '2026-08-24', '2026-08-27', 'Done'),
    ('WBS 8.4', 'Finance Invoices Workspace', 'Categorized invoice tabs (Drafts, Approved, Exported, Failed) with sticky headers and search', 'Abhishek (Primary) / Rahul (Support)', 18.0, '2026-08-25', '2026-08-28', 'Done'),
    ('WBS 8.5', 'Staged Email Inbox UI', 'Document staging workspace, polling trigger button, email metadata viewer, and promotion modal', 'Neeraj / Abhishek', 16.0, '2026-08-26', '2026-08-29', 'Done'),
    ('WBS 8.6', 'Integrations Management Hub', 'Zoho Books connection status, master data browser, and IMAP configuration drawer', 'Abhishek (Primary) / Raju (Support)', 16.0, '2026-08-26', '2026-08-29', 'Done'),

    # 9. Security, Quality Assurance & Deployment
    ('WBS 9.1', 'Authentication & RBAC', 'JWT bearer authentication, role-based access control (ADMIN, FINANCE, VIEWER), and login UI', 'Rahul (Primary) / Abhishek (Support)', 18.0, '2026-08-25', '2026-08-28', 'Done'),
    ('WBS 9.2', 'Multi-Tenancy & Data Isolation', 'Strict tenant_id enforcement across all database queries, migrations, and storage paths', 'Rahul (Primary) / Abhishek (Support)', 16.0, '2026-08-26', '2026-08-29', 'Done'),
    ('WBS 9.3', 'Automated Regression Test Suite', '113 unit, integration, and security test scenarios covering Stages 1 to 6', 'Rahul (Primary) / Abhishek (Support)', 24.0, '2026-08-27', '2026-08-30', 'Done'),
    ('WBS 9.4', 'Vendor Real-time Validation API', 'Live GSTIN / PAN validation against government portals / external APIs', 'Rahul (Primary) / Raju (Support)', 16.0, '2026-08-29', '2026-09-02', 'In Progress'),
    ('WBS 9.5', 'Production Deployment & Cloud Setup', 'Docker containerization, Supabase DB pooling, environment configs, and SSL reverse proxy', 'Rahul (Primary) / Abhishek (Support)', 20.0, '2026-08-29', '2026-09-03', 'In Progress'),
    ('WBS 9.6', 'End-to-End Enterprise UAT Sign-off', 'Finance team walkthrough, edge-case reconciliation resolution, and user sign-off', 'All Team Members', 20.0, '2026-08-30', '2026-09-04', 'Needs Verification')
]

for r_idx, row in enumerate(wbs_data, 4):
    for c_idx, val in enumerate(row, 1):
        cell = ws_wbs.cell(r_idx, c_idx, val)
        cell.border = thin_border
        if c_idx == 1:
            cell.font = font_data_bold
            cell.alignment = align_center
        elif c_idx in (5, 6, 7):
            cell.font = font_data
            cell.alignment = align_center if c_idx != 5 else align_right
        elif c_idx == 8:
            cell.alignment = align_center
            apply_status_style(cell, val)
        else:
            cell.font = font_data
            cell.alignment = align_left
    ws_wbs.row_dimensions[r_idx].height = 20

# Summary Row at the bottom of WBS
r_sum = len(wbs_data) + 4
ws_wbs.cell(r_sum, 4, 'Total Estimated Hours:')
ws_wbs.cell(r_sum, 4).font = font_data_bold
ws_wbs.cell(r_sum, 4).alignment = align_right

ws_wbs.cell(r_sum, 5, f'=SUM(E4:E{r_sum-1})')
ws_wbs.cell(r_sum, 5).font = font_data_bold
ws_wbs.cell(r_sum, 5).alignment = align_right
ws_wbs.cell(r_sum, 5).border = thin_border


# ==========================================
# 3. SHEET: QA Test Suite (43 Cases)
# ==========================================
ws_qa = wb['QA Test Suite (43 Cases)']
ws_qa.views.sheetView[0].showGridLines = True
clear_sheet_contents(ws_qa)

# Title
ws_qa.cell(1, 1, 'QA Test Suite & Functional Verification Matrix (43 Enterprise Scenarios)')
ws_qa.merge_cells('A1:F1')
ws_qa.cell(1, 1).font = font_title
ws_qa.cell(1, 1).fill = fill_title
ws_qa.cell(1, 1).alignment = align_left
ws_qa.row_dimensions[1].height = 35

# Headers
qa_headers = ['Use Case ID', 'Functional Area', 'Scenario Description', 'Responsible Owner', 'Expected System Behavior', 'Verification Status']
for c, h in enumerate(qa_headers, 1):
    cell = ws_qa.cell(3, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center if c in (1, 6) else align_left
    cell.border = thin_border
ws_qa.row_dimensions[3].height = 26

# Full 43 Specific QA Scenarios covering all aspects
qa_cases = [
    ('TC-01', 'Vendor Management', 'Out-of-state vendor with mismatched tax bucket', 'Rahul (Primary) / Abhishek (Support)', 'Auto-detected by GST engine, flagged MISMATCH, blocks premature sync', 'Passed'),
    ('TC-02', 'GST Compliance', 'Intrastate Maharashtra to Maharashtra commercial supply', 'Abhishek (Primary) / Rahul (Support)', 'Auto-resolves CGST 9% + SGST 9% tax accounts deterministically', 'Passed'),
    ('TC-03', 'GST Compliance', 'Special Economic Zone (SEZ) Supplier invoice', 'Rahul (Primary) / Abhishek (Support)', 'Applies SEZ 0% IGST override regardless of physical vendor state', 'Passed'),
    ('TC-04', 'TDS Engine', 'IT / DevOps Technical Consulting Invoice', 'Raju', 'Auto-predicts Section 194J Professional Fees (2% or 10%)', 'Passed'),
    ('TC-05', 'TDS Engine', 'Commercial Lease / Rent Invoice', 'Raju', 'Auto-predicts Section 194I Land/Building Rent (10%)', 'Passed'),
    ('TC-06', 'TDS Engine', 'Contractor Works / Maintenance Services Invoice', 'Raju', 'Auto-predicts Section 194C Contractor Withholding (1% / 2%)', 'Passed'),
    ('TC-07', 'TDS Engine', 'Vendor PAN non-furnishing / invalid PAN detection', 'Raju', 'Triggers higher deduction rate penalty (20%) under Section 206AA', 'Passed'),
    ('TC-08', 'TDS Engine', 'Invoice date in past year vs active TDS financial lock', 'Raju', 'Auto-healer retries with active effective tax date in current fiscal cycle', 'Passed'),
    ('TC-09', 'COA Category Engine', 'IT Hardware purchase (27-inch 4K LED Monitors)', 'Abhishek (Primary) / Rahul (Support)', 'Maps semantically to Computer Equipment asset/expense account', 'Passed'),
    ('TC-10', 'Vendor Profile', 'Vendor update with new Bank Account & IFSC details', 'Raju', 'Zoho vendor contact upserted with company name, email, and bank notes', 'Passed'),
    ('TC-11', 'Financial Validation', 'Math discrepancy > ₹1.00 (Line items taxable sum vs Subtotal)', 'Rahul (Primary) / Abhishek (Support)', 'Flags overall_status = MISMATCH, records diff warnings, marks review needed', 'Passed'),
    ('TC-12', 'Financial Validation', 'Multi-line item arithmetic check (Qty × Unit Price vs Total)', 'Rahul (Primary) / Abhishek (Support)', 'Validates per-line arithmetic independently and isolates faulty line items', 'Passed'),
    ('TC-13', 'Financial Validation', 'Header tax total vs sum of CGST + SGST + IGST components', 'Rahul (Primary) / Abhishek (Support)', 'Confirms exact penny equality between tax header and component sums', 'Passed'),
    ('TC-14', 'Journal Generator', 'Double-entry balance check (Total Debits == Total Credits)', 'Abhishek (Primary) / Rahul (Support)', 'Enforces Debits == Credits within ₹1.00 tolerance or marks UNBALANCED', 'Passed'),
    ('TC-15', 'Journal Generator', 'Ineligible input tax under Section 17(5) accounting', 'Abhishek (Primary) / Raju (Support)', 'Routes ineligible GST debits to Ineligible Input Tax Expense account', 'Passed'),
    ('TC-16', 'Journal Generator', 'TDS deduction impact on Accounts Payable credit', 'Abhishek (Primary) / Raju (Support)', 'Calculates Vendor AP Credit as Gross Total minus Withheld TDS', 'Passed'),
    ('TC-17', 'Journal Generator', 'Round-off penny difference handling (<= ₹1.00)', 'Abhishek (Primary) / Rahul (Support)', 'Automatically injects balancing debit/credit round-off adjustment line', 'Passed'),
    ('TC-18', 'HITL Approval', 'Approve invoice with complete Finance-approved COA lines', 'Rahul (Primary) / Abhishek (Support)', 'Generates authoritative journal, locks invoice, stamps approved_by and approved_at', 'Passed'),
    ('TC-19', 'HITL Approval', 'Attempt approval on unapproved AI-suggested COA line', 'Rahul (Primary) / Abhishek (Support)', 'Throws HTTP 400 rejecting approval until Finance explicitly confirms account', 'Passed'),
    ('TC-20', 'HITL Approval', 'Attempt approval on unbalanced journal entry', 'Rahul (Primary) / Abhishek (Support)', 'Throws HTTP 400 blocking approval with exact Debit/Credit difference details', 'Passed'),
    ('TC-21', 'HITL Rejection', 'Finance manager rejects invoice with mandatory reason note', 'Abhishek (Primary) / Rahul (Support)', 'Sets status to REJECTED, unlocks edits, and logs audit event', 'Passed'),
    ('TC-22', 'Audit Trail', 'Full lifecycle audit tracking from upload to Zoho sync', 'Rahul (Primary) / Abhishek (Support)', 'Records timestamped audit log entries with user email, action, and IP context', 'Passed'),
    ('TC-23', 'Duplicate Detection', 'Exact duplicate invoice file upload (same SHA-256 hash)', 'Rahul (Primary) / Abhishek (Support)', 'Rejects upload with HTTP 409 Conflict and existing invoice link', 'Passed'),
    ('TC-24', 'Duplicate Detection', 'Re-uploading failed / staged document file', 'Abhishek (Primary) / Rahul (Support)', 'Promotes existing staged record to processing without duplicating records', 'Passed'),
    ('TC-25', 'Multi-Tenancy', 'Tenant A accessing Tenant B invoice via direct UUID', 'Rahul (Primary) / Abhishek (Support)', 'Returns HTTP 404/403 strictly isolating tenant data across all queries', 'Passed'),
    ('TC-26', 'Auth & RBAC', 'Viewer role attempting invoice approval or COA override', 'Rahul (Primary) / Abhishek (Support)', 'Returns HTTP 403 Forbidden enforcing ADMIN/FINANCE role requirements', 'Passed'),
    ('TC-27', 'IMAP Email Ingestion', 'Polling Gmail mailbox with Google App Password over SSL', 'Neeraj', 'Fetches UNSEEN emails with PDF/image attachments and stages in inbox', 'Passed'),
    ('TC-28', 'IMAP Email Ingestion', 'Catching IMAP authentication failures with clear user guidance', 'Neeraj', 'Returns actionable error instructing Google 2FA & App Password configuration', 'Passed'),
    ('TC-29', 'Inbox Staging', 'Promoting staged email attachment to main invoice review pipeline', 'Neeraj', 'Creates new invoice record, triggers VLM extraction, updates staged status', 'Passed'),
    ('TC-30', 'PDF Page Rendering', 'Multi-page PDF invoice preview generation via PyMuPDF', 'Neeraj', 'Renders all pages as high-res images for side-by-side review workspace', 'Passed'),
    ('TC-31', 'Zoho OAuth2', 'Initial OAuth2 handshake and authorization code exchange', 'Raju', 'Exchanges authorization code for access/refresh tokens and encrypts in DB', 'Passed'),
    ('TC-32', 'Zoho OAuth2', 'Access token expiration during background operations', 'Raju', 'Auto-refreshes access token transparently using encrypted refresh token', 'Passed'),
    ('TC-33', 'Zoho Master Data', 'Fetching Zoho Chart of Accounts, Taxes, and Payment Terms', 'Raju', 'Synchronizes remote Zoho master data into local cache with summary counts', 'Passed'),
    ('TC-34', 'Zoho Bill Export', 'Exporting approved invoice with valid accounts and tax IDs', 'Raju', 'Creates vendor bill in Zoho Books and updates zoho_bill_id on invoice', 'Passed'),
    ('TC-35', 'Zoho Bill Export', 'Attempting export of unapproved or draft invoice', 'Raju', 'Blocks export with HTTP 400 requiring invoice approval first', 'Passed'),
    ('TC-36', 'Zoho Resilience', 'Zoho API rate limit (429) or transient network error handling', 'Raju', 'Retries with exponential backoff and marks EXPORT_FAILED with reason if exhausted', 'Passed'),
    ('TC-37', 'UI Responsiveness', 'Side-by-side PDF viewer and 7-pillar validation cards', 'Abhishek (Primary) / Rahul (Support)', 'Smooth document zoom/pan alongside responsive data editing forms', 'Passed'),
    ('TC-38', 'UI Resilience', 'Preserving edited form fields across browser tab switching', 'Abhishek (Primary) / Rahul (Support)', 'Maintains local review state and provides explicit Save Draft action', 'Passed'),
    ('TC-39', 'Credit Notes Support', 'Handling vendor credit notes / refund adjustments', 'Raju', 'Maps negative line items and credit note payload structure to Zoho ERP', 'In Progress'),
    ('TC-40', 'Live Vendor GSTIN API', 'Real-time taxpayer verification against GST portal', 'Rahul (Primary) / Raju (Support)', 'Queries government API to verify active registration status and legal name', 'In Progress'),
    ('TC-41', 'Closed Accounting Period', 'Posting invoice dated in locked prior financial period', 'Rahul (Primary) / Abhishek (Support)', 'Enforces closed period lock and suggests adjustment in open fiscal period', 'Partially Completed'),
    ('TC-42', 'Foreign Currency (FX)', 'Multi-currency invoice with exchange rate conversion to INR', 'Rahul (Primary) / Abhishek (Support)', 'Extracts currency code and converts base amounts to INR for GL journal', 'In Progress'),
    ('TC-43', 'End-to-End Enterprise UAT', 'Full multi-vendor batch invoice processing and audit sign-off', 'All Team Members', 'Finance user acceptance testing across complete 7-pillar lifecycle', 'Needs Verification')
]

for r_idx, row in enumerate(qa_cases, 4):
    for c_idx, val in enumerate(row, 1):
        cell = ws_qa.cell(r_idx, c_idx, val)
        cell.border = thin_border
        if c_idx == 1:
            cell.font = font_data_bold
            cell.alignment = align_center
        elif c_idx == 6:
            cell.alignment = align_center
            apply_status_style(cell, val)
        elif c_idx == 4:
            cell.font = font_data_bold
            cell.alignment = align_left
        else:
            cell.font = font_data
            cell.alignment = align_left
    ws_qa.row_dimensions[r_idx].height = 20


# ==========================================
# 4. SHEET: Timeline & Weekly Schedule
# ==========================================
ws_time = wb['Timeline & Weekly Schedule']
ws_time.views.sheetView[0].showGridLines = True
clear_sheet_contents(ws_time)

# Title
ws_time.cell(1, 1, 'Project Timeline & Phase Execution Schedule — Actual Repository Progress')
ws_time.merge_cells('A1:E1')
ws_time.cell(1, 1).font = font_title
ws_time.cell(1, 1).fill = fill_title
ws_time.cell(1, 1).alignment = align_left
ws_time.row_dimensions[1].height = 35

# Section 1 Header
ws_time.cell(3, 1, 'PHASE & SPRINT EXECUTION BREAKDOWN')
ws_time.cell(3, 1).font = font_section

time_headers = ['Phase / Sprint', 'Core Task Deliverables', 'Start Date', 'End Date', 'Actual Status']
for c, h in enumerate(time_headers, 1):
    cell = ws_time.cell(4, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center if c in (3, 4, 5) else align_left
    cell.border = thin_border
ws_time.row_dimensions[4].height = 24

timeline_data = [
    ('Phase 1: Architecture & Ingestion', 'Multi-tenant DB schema, storage backend, hash deduplication, Qwen-VL integration', '2026-08-01', '2026-08-08', 'Done'),
    ('Phase 2: Compliance Engines', 'Deterministic GST, ITC Section 17(5), TDS Withholding (194C/J/I/Q) engines', '2026-08-08', '2026-08-16', 'Done'),
    ('Phase 3: Financial Validation & GL', 'Stage 5 math reconciler, Stage 6 double-entry journal generator, relational GL sync', '2026-08-16', '2026-08-24', 'Done'),
    ('Phase 4: Integrations & Review UI', 'Zoho Books OAuth2 & bill export, IMAP email ingestion, 7-pillar review hub', '2026-08-20', '2026-08-28', 'Done'),
    ('Phase 5: Automated Testing & Security', '113 automated unit/integration tests, RBAC security, multi-tenant isolation', '2026-08-25', '2026-08-30', 'Done'),
    ('Phase 6: UAT & Production Launch', 'End-to-end enterprise UAT walkthrough, edge-case tuning, production deployment', '2026-08-30', '2026-09-04', 'In Progress')
]

for r_idx, row in enumerate(timeline_data, 5):
    for c_idx, val in enumerate(row, 1):
        cell = ws_time.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx == 1 else font_data
        cell.border = thin_border
        cell.alignment = align_center if c_idx in (3, 4, 5) else align_left
        if c_idx == 5:
            apply_status_style(cell, val)
    ws_time.row_dimensions[r_idx].height = 20

# Section 2: Team Member Sprint Deliverables Summary
r_start_ws = 13
ws_time.cell(r_start_ws, 1, 'WEEKLY TEAM CONTRIBUTION & SPRINT SUMMARY')
ws_time.cell(r_start_ws, 1).font = font_section

sprint_headers = ['Sprint Week', 'Key Focus Areas', 'Primary Contributors', 'Key Deliverables & Milestones']
for c, h in enumerate(sprint_headers, 1):
    cell = ws_time.cell(r_start_ws + 1, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_left
    cell.border = thin_border
ws_time.row_dimensions[r_start_ws + 1].height = 24

sprint_rows = [
    ('Week 1 (Aug 01 - Aug 07)', 'Backend Foundations, DB Schemas, Storage, Dual-State JSONB, Qwen-VL Parsing', 'Rahul, Abhishek', 'Operational backend API, invoice upload pipeline, Pydantic schemas'),
    ('Week 2 (Aug 08 - Aug 15)', 'GST Engine, ITC Classifier, TDS Tax Brain, Semantic COA Categorization', 'Raju, Rahul, Abhishek', 'Deterministic tax rules, PAN entity slicer, Section 194C/J/I rules'),
    ('Week 3 (Aug 16 - Aug 23)', 'Financial Validation, Balanced GL Generator, Zoho OAuth2, IMAP Email Ingestion', 'Rahul, Abhishek, Raju, Neeraj', 'Stage 5 math reconciler, Stage 6 double entry journal, Zoho client, IMAP service'),
    ('Week 4 (Aug 24 - Aug 30)', '7-Pillar Review Hub UI, Staged Inbox Workspace, GL Inspector, Zoho Bill Export', 'Abhishek, Rahul, Raju, Neeraj', 'Next.js review workspace, Zoho bill sync, 111 passed automated tests'),
    ('Week 5 (Aug 31 - Sep 06)', 'Enterprise UAT Execution, Live Edge-Case Fixes, Production Cloud Deployment', 'All Team Members', 'Final UAT approval, production readiness, and live customer handover')
]

for r_idx, row in enumerate(sprint_rows, r_start_ws + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_time.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx == 1 else font_data
        cell.border = thin_border
        cell.alignment = align_left
    ws_time.row_dimensions[r_idx].height = 22


# ==========================================
# 5. SHEET: RACI Matrix
# ==========================================
ws_raci = wb['RACI Matrix']
ws_raci.views.sheetView[0].showGridLines = True
clear_sheet_contents(ws_raci)

# Title
ws_raci.cell(1, 1, 'Resource Responsibility Matrix (RACI) — Real Team Structure')
ws_raci.merge_cells('A1:E1')
ws_raci.cell(1, 1).font = font_title
ws_raci.cell(1, 1).fill = fill_title
ws_raci.cell(1, 1).alignment = align_left
ws_raci.row_dimensions[1].height = 35

ws_raci.cell(3, 1, 'RACI Legend:  R = Responsible (Did the Work)  |  A = Accountable (Approver/Owner)  |  C = Consulted  |  I = Informed')
ws_raci.cell(3, 1).font = Font(name='Calibri', size=10, italic=True, bold=True, color='595959')

# RACI Headers
raci_headers = ['Functional Module / Workstream', 'Rahul (Backend / Arch Lead)', 'Abhishek (Frontend / Core Co-Lead)', 'Raju (TDS / Zoho ERP Lead)', 'Neeraj (Email / IMAP Lead)']
for c, h in enumerate(raci_headers, 1):
    cell = ws_raci.cell(5, c, h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center if c > 1 else align_left
    cell.border = thin_border
ws_raci.row_dimensions[5].height = 26

raci_data = [
    ('Overall System Architecture & Multi-Tenant Database', 'A / R', 'C / R', 'C', 'I'),
    ('Invoice Ingestion, File Storage & Hash Deduplication', 'A / R', 'C', 'I', 'I'),
    ('IMAP Email Auto-Ingestion & Inbox Staging Service', 'I', 'C', 'I', 'A / R'),
    ('PDF Multi-Page Renderer & Document Previews', 'I', 'C', 'I', 'A / R'),
    ('Qwen3-VL & Qwen3-4B AI Model Integrations & Schemas', 'A / R', 'C / R', 'I', 'I'),
    ('Deterministic GST Engine & State Supply Type Rules', 'C', 'A / R', 'C', 'I'),
    ('ITC Eligibility & Section 17(5) Blocked Credit Engine', 'C', 'A / R', 'C', 'I'),
    ('TDS Withholding Engine (Sections 194C, 194J, 194I, 194Q)', 'I', 'C', 'A / R', 'I'),
    ('Vendor PAN Validation & Entity Slicing (206AA)', 'I', 'C', 'A / R', 'I'),
    ('Stage 5 Financial Validation & Mathematical Cross-Reconciliation', 'A / R', 'C', 'I', 'I'),
    ('Stage 6 Double-Entry Journal Generator & GL Tables Sync', 'C', 'A / R', 'C', 'I'),
    ('HITL Review API, Authoritative Approval & Rejection Lock', 'A / R', 'C / R', 'I', 'I'),
    ('Audit Trail Logging & Event Capture Service', 'A / R', 'C', 'I', 'I'),
    ('Zoho Books OAuth2 Resilient Client & Auto-Healer', 'I', 'I', 'A / R', 'I'),
    ('Zoho Master Data Sync (COA, Taxes, Vendors, Terms)', 'I', 'I', 'A / R', 'I'),
    ('Zoho Bill Creation & Multi-Account GL Export Engine', 'I', 'C', 'A / R', 'I'),
    ('Next.js Frontend Architecture & AppShell Navigation', 'C', 'A / R', 'I', 'I'),
    ('7-Pillar Review Hub UI & Split-Screen Document Viewer', 'C', 'A / R', 'I', 'I'),
    ('GL Double-Entry Inspector UI & Real-Time Balance Badge', 'C', 'A / R', 'I', 'I'),
    ('Integrations Hub (Zoho Status & IMAP Settings Drawer)', 'I', 'A / R', 'C', 'C'),
    ('Security (JWT Auth, RBAC Roles, Tenant Isolation)', 'A / R', 'C', 'I', 'I'),
    ('Automated Testing Suite (Unit, Integration, Security)', 'A / R', 'C / R', 'C', 'C'),
    ('Finance Team UAT Walkthrough & Production Deployment', 'A / R', 'A / R', 'A / R', 'A / R')
]

for r_idx, row in enumerate(raci_data, 6):
    for c_idx, val in enumerate(row, 1):
        cell = ws_raci.cell(r_idx, c_idx, val)
        cell.font = font_data_bold if c_idx == 1 else font_data
        cell.border = thin_border
        cell.alignment = align_center if c_idx > 1 else align_left
        # Color coding for RACI cells
        if c_idx > 1:
            if 'A / R' in val:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color='006100')
            elif val == 'A':
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color='203764')
            elif val == 'R':
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color='375623')
            elif 'C' in val:
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=False, color='7F6000')
            elif val == 'I':
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=False, color='595959')
    ws_raci.row_dimensions[r_idx].height = 20


# Set appropriate column widths across all sheets
col_widths = {
    'Executive Summary': {'A': 20.0, 'B': 45.0, 'C': 40.0, 'D': 30.0, 'E': 25.0},
    'Master Task List (WBS)': {'A': 12.0, 'B': 30.0, 'C': 58.0, 'D': 35.0, 'E': 12.0, 'F': 14.0, 'G': 14.0, 'H': 18.0},
    'QA Test Suite (43 Cases)': {'A': 14.0, 'B': 24.0, 'C': 48.0, 'D': 35.0, 'E': 55.0, 'F': 18.0},
    'Timeline & Weekly Schedule': {'A': 30.0, 'B': 58.0, 'C': 25.0, 'D': 45.0, 'E': 18.0},
    'RACI Matrix': {'A': 48.0, 'B': 26.0, 'C': 28.0, 'D': 26.0, 'E': 26.0}
}

for sname, widths in col_widths.items():
    ws = wb[sname]
    for c_letter, w in widths.items():
        ws.column_dimensions[c_letter].width = w

# Save updated workbook
wb.save(dst_file)
print(f'Successfully updated and saved master tracker to: {dst_file}')
